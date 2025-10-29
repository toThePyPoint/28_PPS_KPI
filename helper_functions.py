import logging

from datetime import datetime, date, timedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Alignment


def copy_row_format(ws, source_row, target_row):
    """
    Copies border formatting from source_row to target_row in the given worksheet.

    :param ws: Worksheet object
    :param source_row: int: Row number to copy formatting from
    :param target_row: int: Row number to apply formatting to
    """
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        # Create a new Border object by copying the properties of the source cell
        if source_cell.border:
            new_border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
            target_cell.border = new_border  # Apply new border

        # Copy Text Wrapping
        if source_cell.alignment:
            new_alignment = Alignment(
                wrap_text=source_cell.alignment.wrap_text  # Preserve wrapping
            )
            target_cell.alignment = new_alignment


def append_data_to_excel(status_file, data_dict, error_path, sheet_name):
    """
    Appends a new row to the given sheet in the given Excel file using the status_dict.

    :param sheet_name: sheet_name of excel status file
    :param error_path: path to error file
    :param status_file: str: Path to the Excel file
    :param data_dict: dict: Dictionary containing status messages
    """
    logging.basicConfig(
        filename=error_path,
        level=logging.ERROR,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    try:
        # Load the existing workbook
        wb = load_workbook(status_file)

        # Select the "MRP_STOCKS" sheet
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
            return

        ws = wb[sheet_name]

        # Get headers from the first row
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        # Find the first empty row within the framed area
        first_empty_row = None
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip headers
            if all(ws.cell(row=row, column=col).value in [None, ""] for col in range(1, ws.max_column + 1)):
                first_empty_row = row
                break

        # If no empty row is found, insert at the end
        if first_empty_row is None:
            first_empty_row = ws.max_row + 1

        # Insert new data at the first empty row
        ws.insert_rows(first_empty_row)

        # Copy border formatting from the row above
        if first_empty_row > 2:  # Ensure it's not the header row
            copy_row_format(ws, first_empty_row - 1, first_empty_row)

        today = datetime.today()
        today_str = today.strftime('%Y-%m-%d')

        # Add date in column A
        ws.cell(row=first_empty_row, column=1, value=today_str)

        # Fill in values based on dictionary keys matching headers
        for col, header in enumerate(headers[1:], start=2):  # Start from column 2 (B) as A is for timestamp
            ws.cell(row=first_empty_row, column=col, value=str(data_dict.get(header, "")))

        # Append the row and save the file
        wb.save(status_file)

        print(f"KPIs updated successfully - {data_dict['LINE']}!")

    except Exception as e:
        logging.error("Error occurred", exc_info=True)
        print("Error occurred: ", e)
        print(f"Check {error_path} file for details")


def get_nth_working_day(num_of_days: int) -> pd.Timestamp:
    """
    Returns the n-th working day (as pandas.Timestamp)
    starting from the next working day.
    """
    current_date = date.today() + timedelta(days=1)  # start from tomorrow

    while num_of_days > 0:
        if current_date.weekday() < 5:  # Mon-Fri = 0-4
            num_of_days -= 1
            if num_of_days == 0:
                break
        current_date += timedelta(days=1)

    return pd.to_datetime(current_date)  # ensures pandas-compatible Timestamp
